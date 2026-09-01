"""Parse KASHMIR_OIL_SALES Excel into baPerformance.generated.json"""
from __future__ import annotations

import json
import re
from collections import defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / 'KASHMIR_OIL_SALES_27_AUGUST.xlsx'
OUT = ROOT / 'src' / 'data' / 'baPerformance.generated.json'

MONTH_MAP = {
    'JAN': 'January',
    'FEB': 'February',
    'MARCH': 'March',
    'APRIL': 'April',
    'MAY': 'May',
    'JUNE': 'June',
    'JULY': 'July',
    'AUGUST': 'August',
    'SEP': 'September',
    'SEPT': 'September',
    'OCT': 'October',
    'NOV': 'November',
    'DEC': 'December',
}

# Towns present in source Excel but outside the Kashmir programme scope
EXCLUDED_TOWNS = {'Daska', 'Muridke'}

SKU_COLUMNS = [
    'Pouch 1LTR',
    'POUCH 1KG',
    'BKT 2.5KG',
    'BKT 5KG',
    'SUP 1LTR',
    'BTL 3LTR',
    'BTL 4.5LTR',
    'CAN 10LTR',
    'TIN 5LTR',
    'POUCH 1X5 KG Box',
    'BKT 10KG',
    'POUCH 1 KG',
    'BUCKET 5 KG',
]


def sheet_month(name: str) -> str | None:
    m = re.search(r'-(JAN|FEB|MARCH|APRIL|MAY|JUNE|JULY|AUGUST|SEP|SEPT|OCT|NOV|DEC)$', name, re.I)
    if not m:
        return None
    return MONTH_MAP.get(m.group(1).upper())


def included_town(town: str) -> bool:
    return town.strip() not in EXCLUDED_TOWNS


def num(v) -> float:
    if v is None or v == '' or str(v).startswith('#'):
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def parse_summary(wb) -> dict[tuple[str, str], dict]:
    """Town-month aggregates from SUMMARY-* sheets."""
    town_month: dict[tuple[str, str], dict] = {}
    for sheet in wb.sheetnames:
        month = sheet_month(sheet)
        if not sheet.startswith('SUMMARY-') or not month:
            continue
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        header_idx = None
        headers: list[str] = []
        for i, row in enumerate(rows[:6]):
            if row and row[0] == 'S#':
                header_idx = i
                headers = [str(c).strip() if c is not None else '' for c in row]
                break
        if header_idx is None:
            continue

        def col(name: str) -> int | None:
            for i, h in enumerate(headers):
                if name.lower() in h.lower():
                    return i
            return None

        i_town = col('TOWN') or 1
        i_int = col('TOTAL INTERCEPTIONS') or col('TOTAL INTERCEPT')
        i_prod = col('Productive Calls') or col('PRODUCTIVE')
        i_oil = col('SALES-OIL')
        i_ghee = col('SALES-GHEE')
        i_waadi = col('SALES-WAADI')
        i_target = col('TARGET')
        i_sales = col('TOTAL SALES')

        for row in rows[header_idx + 1 :]:
            if not row or row[0] is None:
                continue
            town = row[i_town]
            if not town or not isinstance(town, str):
                continue
            town = town.strip()
            if not included_town(town):
                continue
            intercepted = num(row[i_int]) if i_int is not None else 0
            productive = num(row[i_prod]) if i_prod is not None else 0
            oil = num(row[i_oil]) if i_oil is not None else 0
            ghee = num(row[i_ghee]) if i_ghee is not None else 0
            waadi = num(row[i_waadi]) if i_waadi is not None else 0
            target = num(row[i_target]) if i_target is not None else 0
            sales = num(row[i_sales]) if i_sales is not None else oil + ghee + waadi
            town_month[(town, month)] = {
                'customersIntercepted': round(intercepted),
                'productiveCalls': round(productive),
                'targetLtrKg': round(target),
                'salesLtrKg': round(sales, 1),
                'oilSales': round(oil, 1),
                'gheeSales': round(ghee, 1),
                'waadiSales': round(waadi, 1),
            }
    return town_month


def parse_pivot(wb) -> dict[tuple[str, str, str], dict]:
    """Store-month records from PIVOT-* sheets (right-hand sales table)."""
    records: dict[tuple[str, str, str], dict] = {}
    for sheet in wb.sheetnames:
        month = sheet_month(sheet)
        if not sheet.startswith('PIVOT-') or not month:
            continue
        ws = wb[sheet]
        start_col = 0
        for row in ws.iter_rows(max_row=6, values_only=True):
            vals = list(row)
            for i, v in enumerate(vals):
                if v == 'Sum of Productive Calls':
                    start_col = i - 2  # town at i-2, store at i-1
                    break
            if start_col > 0:
                break
        if start_col <= 0:
            start_col = 7  # fallback for standard layout

        for row in ws.iter_rows(min_row=5, values_only=True):
            vals = list(row)
            if len(vals) <= start_col + 6:
                continue
            town = vals[start_col]
            store = vals[start_col + 1]
            if not town or not store or not isinstance(town, str) or not isinstance(store, str):
                continue
            if store == 'Grand Total' or town == 'Grand Total':
                continue
            town = town.strip()
            if not included_town(town):
                continue
            productive = num(vals[start_col + 2])
            non_prod = num(vals[start_col + 3])
            oil = num(vals[start_col + 4])
            ghee = num(vals[start_col + 5])
            waadi = num(vals[start_col + 6])
            intercepted = productive + non_prod
            sales = oil + ghee + waadi
            key = (town, month, store.strip())
            records[key] = {
                'customersIntercepted': round(intercepted),
                'productiveCalls': round(productive),
                'salesLtrKg': round(sales, 1),
                'oilSales': round(oil, 1),
                'gheeSales': round(ghee, 1),
                'waadiSales': round(waadi, 1),
            }
    return records


def parse_sales(wb) -> dict[tuple[str, str, str], dict]:
    """Week + SKU aggregates from SALES-* sheets."""
    agg: dict[tuple[str, str, str], dict] = defaultdict(
        lambda: {'weekSales': defaultdict(float), 'skuSales': defaultdict(float), 'target': 0.0}
    )

    for sheet in wb.sheetnames:
        month = sheet_month(sheet)
        if not sheet.startswith('SALES-') or not month:
            continue
        ws = wb[sheet]
        headers: list[str] | None = None
        for row in ws.iter_rows(max_row=6, values_only=True):
            if row and row[0] == 'ATTENDANCE' and row[6] == 'TOWN':
                headers = [str(c).strip() if c is not None else '' for c in row]
                break
        if not headers:
            continue

        def idx(name: str) -> int | None:
            for i, h in enumerate(headers):
                if h.lower() == name.lower():
                    return i
            return None

        i_town = idx('TOWN')
        i_store = idx('STORE')
        i_week = idx('WEEKS')
        i_total = idx('TOTAL SALES (LTR/KG)')
        if i_town is None or i_store is None:
            continue

        sku_idx = {sku: idx(sku) for sku in SKU_COLUMNS if idx(sku) is not None}

        for row in ws.iter_rows(min_row=5, values_only=True):
            vals = list(row)
            if len(vals) <= max(i_town, i_store):
                continue
            town, store = vals[i_town], vals[i_store]
            if not town or not store:
                continue
            town = str(town).strip()
            if not included_town(town):
                continue
            key = (town, month, str(store).strip())
            if i_week is not None and vals[i_week] is not None:
                week = int(num(vals[i_week]))
                total = num(vals[i_total]) if i_total is not None else 0
                agg[key]['weekSales'][week] += total
            for sku, si in sku_idx.items():
                v = num(vals[si])
                if v:
                    agg[key]['skuSales'][sku] += v
    return agg


def main() -> None:
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    town_month = parse_summary(wb)
    pivot = parse_pivot(wb)
    sales_agg = parse_sales(wb)

    records = []
    towns = set()
    months = set()
    stores_by_town: dict[str, set[str]] = defaultdict(set)

    keys = set(pivot.keys()) | set(sales_agg.keys())
    for (town, month, store) in keys:
        if not included_town(town):
            continue
        towns.add(town)
        months.add(month)
        stores_by_town[town].add(store)

        base = pivot.get((town, month, store), {})
        tm = town_month.get((town, month), {})
        sa = sales_agg.get((town, month, store), {})

        oil = base.get('oilSales', 0)
        ghee = base.get('gheeSales', 0)
        waadi = base.get('waadiSales', 0)
        sales = base.get('salesLtrKg', oil + ghee + waadi)

        # Allocate town target to stores by sales share
        town_target = tm.get('targetLtrKg', 0)
        town_sales = tm.get('salesLtrKg', 0) or 1
        store_target = round(town_target * (sales / town_sales), 1) if sales else 0

        week_sales = [
            {'week': w, 'sales': round(v, 1)}
            for w, v in sorted(sa.get('weekSales', {}).items())
        ]
        sku_sales = [
            {'sku': sku, 'sales': round(v, 1)}
            for sku, v in sorted(sa.get('skuSales', {}).items(), key=lambda x: -x[1])
            if v > 0
        ]

        intercepted = base.get('customersIntercepted', 0)
        productive = base.get('productiveCalls', 0)

        records.append(
            {
                'town': town,
                'month': month,
                'store': store,
                'customersIntercepted': intercepted,
                'productiveCalls': productive,
                'targetLtrKg': store_target,
                'salesLtrKg': round(sales, 1),
                'oilSales': round(oil, 1),
                'gheeSales': round(ghee, 1),
                'waadiSales': round(waadi, 1),
                'weekSales': week_sales,
                'skuSales': sku_sales[:10],
            }
        )

    # Town-only fallback records when pivot missing but summary exists
    for (town, month), tm in town_month.items():
        if not included_town(town):
            continue
        if any(r['town'] == town and r['month'] == month for r in records):
            continue
        records.append(
            {
                'town': town,
                'month': month,
                'store': '__ALL__',
                'customersIntercepted': tm['customersIntercepted'],
                'productiveCalls': tm['productiveCalls'],
                'targetLtrKg': tm['targetLtrKg'],
                'salesLtrKg': tm['salesLtrKg'],
                'oilSales': tm['oilSales'],
                'gheeSales': tm['gheeSales'],
                'waadiSales': tm['waadiSales'],
                'weekSales': [],
                'skuSales': [],
            }
        )

    month_order = [
        'January',
        'February',
        'March',
        'April',
        'May',
        'June',
        'July',
        'August',
        'September',
        'October',
        'November',
        'December',
    ]
    sorted_months = [m for m in month_order if m in months]

    payload = {
        'towns': sorted(towns),
        'months': sorted_months,
        'storesByTown': {t: sorted(s) for t, s in stores_by_town.items()},
        'records': records,
    }

    OUT.write_text(json.dumps(payload, indent=2), encoding='utf-8')
    wb.close()
    print(f'Wrote {len(records)} records -> {OUT}')


if __name__ == '__main__':
    main()
